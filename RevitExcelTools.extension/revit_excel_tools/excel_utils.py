"""
Excel utilities for importing and exporting data
"""
try:
    # Try to import openpyxl (recommended)
    from openpyxl import Workbook, load_workbook
    EXCEL_LIBRARY = "openpyxl"
except ImportError:
    try:
        # Fallback to xlrd/xlwt
        import xlrd
        import xlwt
        EXCEL_LIBRARY = "xlrd_xlwt"
    except ImportError:
        # Last resort: use COM (slower but always available on Windows)
        EXCEL_LIBRARY = "com"


class ExcelExporter(object):
    """Export Revit data to Excel"""

    def __init__(self, file_path):
        self.file_path = file_path

    def export_data(self, data):
        """
        Export data to Excel file

        Args:
            data: List of dictionaries containing element data
        """
        if EXCEL_LIBRARY == "openpyxl":
            self._export_openpyxl(data)
        elif EXCEL_LIBRARY == "xlrd_xlwt":
            self._export_xlwt(data)
        else:
            self._export_com(data)

    def _export_openpyxl(self, data):
        """Export using openpyxl library"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "RevitData"

            if not data:
                print("No data to export")
                return

            # Write headers
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Write data rows
            for row_idx, item in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = item.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(self.file_path)
            print(f"Data exported to {self.file_path}")

        except Exception as e:
            print(f"Error exporting with openpyxl: {e}")
            raise

    def _export_xlwt(self, data):
        """Export using xlwt library"""
        try:
            import xlwt
            wb = xlwt.Workbook()
            ws = wb.add_sheet("RevitData")

            if not data:
                print("No data to export")
                return

            # Write headers
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers):
                ws.write(0, col_idx, header)

            # Write data rows
            for row_idx, item in enumerate(data, start=1):
                for col_idx, header in enumerate(headers):
                    value = item.get(header, "")
                    ws.write(row_idx, col_idx, value)

            wb.save(self.file_path)
            print(f"Data exported to {self.file_path}")

        except Exception as e:
            print(f"Error exporting with xlwt: {e}")
            raise

    def _export_com(self, data):
        """Export using COM (Excel COM interface)"""
        try:
            import System
            excel = System.Runtime.InteropServices.Marshal.GetActiveObject("Excel.Application")
            wb = excel.Workbooks.Add()
            ws = wb.Worksheets(1)

            if not data:
                print("No data to export")
                return

            # Write headers
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                ws.Cells(1, col_idx).Value = header

            # Write data rows
            for row_idx, item in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = item.get(header, "")
                    ws.Cells(row_idx, col_idx).Value = value

            wb.SaveAs(self.file_path)
            print(f"Data exported to {self.file_path}")

        except Exception as e:
            print(f"Error exporting with COM: {e}")
            raise


class ExcelImporter(object):
    """Import data from Excel"""

    def __init__(self, file_path):
        self.file_path = file_path

    def import_data(self):
        """
        Import data from Excel file

        Returns:
            List of dictionaries containing row data
        """
        if EXCEL_LIBRARY == "openpyxl":
            return self._import_openpyxl()
        elif EXCEL_LIBRARY == "xlrd_xlwt":
            return self._import_xlrd()
        else:
            return self._import_com()

    def _import_openpyxl(self):
        """Import using openpyxl library"""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            data = []
            headers = None

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    headers = row
                    continue

                if headers:
                    row_dict = {}
                    for col_idx, header in enumerate(headers):
                        if header:
                            value = row[col_idx] if col_idx < len(row) else None
                            row_dict[header] = value
                    if any(row_dict.values()):  # Only add non-empty rows
                        data.append(row_dict)

            print(f"Imported {len(data)} rows from Excel")
            return data

        except Exception as e:
            print(f"Error importing with openpyxl: {e}")
            raise

    def _import_xlrd(self):
        """Import using xlrd library"""
        try:
            import xlrd
            wb = xlrd.open_workbook(self.file_path)
            ws = wb.sheet_by_index(0)

            data = []
            headers = None

            for row_idx in range(ws.nrows):
                row = ws.row_values(row_idx)

                if row_idx == 0:
                    headers = row
                    continue

                if headers:
                    row_dict = {}
                    for col_idx, header in enumerate(headers):
                        if header:
                            value = row[col_idx] if col_idx < len(row) else None
                            row_dict[header] = value
                    if any(row_dict.values()):
                        data.append(row_dict)

            print(f"Imported {len(data)} rows from Excel")
            return data

        except Exception as e:
            print(f"Error importing with xlrd: {e}")
            raise

    def _import_com(self):
        """Import using COM (Excel COM interface)"""
        try:
            import System
            excel = System.Runtime.InteropServices.Marshal.GetActiveObject("Excel.Application")
            wb = excel.Workbooks.Open(self.file_path)
            ws = wb.Worksheets(1)

            data = []
            headers = None
            row_idx = 1

            while True:
                cell_value = ws.Cells(row_idx, 1).Value
                if cell_value is None:
                    break

                if row_idx == 1:
                    headers = []
                    col_idx = 1
                    while True:
                        header = ws.Cells(1, col_idx).Value
                        if header is None:
                            break
                        headers.append(header)
                        col_idx += 1
                else:
                    row_dict = {}
                    for col_idx, header in enumerate(headers, start=1):
                        value = ws.Cells(row_idx, col_idx).Value
                        row_dict[header] = value
                    data.append(row_dict)

                row_idx += 1

            print(f"Imported {len(data)} rows from Excel")
            return data

        except Exception as e:
            print(f"Error importing with COM: {e}")
            raise
